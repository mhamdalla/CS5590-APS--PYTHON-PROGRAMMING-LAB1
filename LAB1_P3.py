import sys
import pandas as pd
from openpyxl import load_workbook      
class Book:
      def __init__(self,listofbooks):#this init method is the first method to be invoked when you create an object
            #what attributes does a library in general have? - for now, let's abstract and just say it has availablebooks (we're not going to program the shelves, and walls in!)
            self.availablebooks=listofbooks
            
      def displayAvailablebooks(self):
                   print("The books we have in our library are as follows:")
                   print("================================")
                   for book in self.availablebooks:
                         print(book)
      def lendBook(self,requestBook):
            requestBook = requestedbook
            if requestBook in self.availablebooks:
                  print("The book you requested has now been borrowed")
                  self.availablebooks.remove(requestBook)
            else:
                  print("Sorry the book you have requested is currently not in the library")
                  self.displayAvailablebooks()      
      def update(self):
          if choice==2:
              if person==1:
                  wb=load_workbook('STUDENT.xlsx')
              elif person==2:
                  wb=load_workbook('FACULTY.xlsx')
                  # select demo.xlsx
              sheet=wb.active
                  # set value for cell B2=2
              sheet.cell(row=index[0], column=3).value = requestedbook 
                  # save workbook 
              if person==1:
                wb.save('STUDENT.xlsx')
              if person==2:
                wb.save('FACULTY.xlsx')
          if choice==3:
              if person==1:
                  wb=load_workbook('STUDENT.xlsx')
              elif person==2:
                  wb=load_workbook('FACULTY.xlsx')
                  # select demo.xlsx
              sheet=wb.active
                  # set value for cell B2=2
              sheet.cell(row=index[0], column=4).value = 'YES' 
                  # save workbook 
              if person==1:
                wb.save('STUDENT.xlsx')
              if person==2:
                wb.save('FACULTY.xlsx')      
      def addBook(self,returnedBook):
            returnedBook= returnedbook
            self.availablebooks.append(returnedBook)
            print("Thanks for returning your borrowed book")
            
class Student:
      def ID(self):
                Id=int(input("Your ID:"))
                df = pd.read_excel('STUDENT.xlsx', sheet_name='Sheet1')
                id_ = df['ID']
                names = dict(df['NAME'])
                if (Id in id_):
                    global index
                    index = [i+2 for i in range(len(id_)) if id_[i] == Id]
                    name = names.get(index[0]-2)
                    print('Welcome',name)
                else:
                    print("Sorry We Couldn't Find your ID")
                    sys.exit()
      def requestBook(self):
            print("Enter the name of the book you'd like to borrow>>")
            global requestedbook
            requestedbook=input()
      def returnBook(self):
            df = pd.read_excel('STUDENT.xlsx', sheet_name='Sheet1')
            rentedbook = df['RENTED BOOK']
            bookname = rentedbook.get(index[0]-2)
            print("You already borrowed",bookname,"Would you like to retun it?")
            g=input() 
            global returnedbook
            if 'y' in g:
                returnedbook=bookname
            else:
                print("Enter the name of the book you'd like to return>>")
                returnedbook=input() 
class Faculty:
    def ID(self):
                Id=int(input("Your ID:"))
                df = pd.read_excel('FACULTY.xlsx', sheet_name='Sheet1')
                id_ = df['ID']
                names = dict(df['NAME'])
                if (Id in id_):
                    global index
                    index = [i+2 for i in range(len(id_)) if id_[i] == Id]
                    name = names.get(index[0]-2)
                    print('Welcome',name)
                else:
                    print("Sorry We Couldn't Find your ID")
                    sys.exit()
    def requestBook(self):
            print("Enter the name of the book you'd like to borrow>>")
            global requestedbook
            requestedbook=input()

    def returnBook(self):
            df = pd.read_excel('STUDENT.xlsx', sheet_name='Sheet1')
            rentedbook = df['RENTED BOOK']
            bookname = rentedbook.get(index[0]-2)
            print("You already borrowed",bookname,"Would you like to retun it?")
            g=input() 
            global returnedbook
            if 'y' in g:
                returnedbook=bookname
            else:
                print("Enter the name of the book you'd like to return>>")
                returnedbook=input()     
           
book=Book(["Python","machine learning","Deep learning","Electromagnetics","math1","electrical circuit"])
student=Student()
faculty=Faculty()
done=False
while done==False:
    print(""" ======LIBRARY MENU=======
                  1. Display all available books
                  2. Request a book
                  3. Return a book
                  4. Exit
                  """)
    global choice
    choice=int(input("Enter Choice:"))
    if choice==1:
        book.displayAvailablebooks()
    elif choice==2:
        print(""" ======Person=======
                             1. student
                             2. faculty
                             """)
        global person
        person=int(input("Enter Choice:"))
        if person==1:
            student.ID()  
            book.lendBook(student.requestBook())
            book.update()
        elif person==2:
            faculty.ID()  
            book.lendBook(faculty.requestBook())
            book.update()
    elif choice==3:
        print(""" ======Person=======
                              1. student
                              2. faculty
                              """)
        person=int(input("Enter Choice:"))
        if person==1:
            student.ID()  
            book.addBook(student.returnBook())
            book.update()
        elif person==2:
            faculty.ID()  
            book.addBook(faculty.returnBook())
            book.update()
    elif choice==4:
        sys.exit()
                 
